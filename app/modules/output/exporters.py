"""
Module 3 — Export Functions
Converts IndicatorResult ORM rows to JSON, CSV, and Excel (XLSX) formats.
Provides three CSV/Excel variants:
  1. 9-column RDTII spec (§2.1) — legacy flat export
  2. RDTII template format — matches the official RDTII template exactly,
     with pillar-section header rows and split reference URL columns (max 5).
  3. Hackathon submission spec (§17) — Economy, Law Name, Article/Section, Discovery Tag, etc.
"""
import io
import json
from typing import Any

import pandas as pd


# ─── Pillar Names (RDTII 2.1 official labels) ────────────────────────────────
PILLAR_NAMES: dict[int, str] = {
    1:  "Tariffs and Trade Defence",
    2:  "Public Procurement",
    3:  "Foreign Direct Investment",
    4:  "Intellectual Property Rights",
    5:  "Telecom Regulations & Competition",
    6:  "Cross-border Data Policies",
    7:  "Domestic Data Privacy & Protection",
    8:  "Internet Intermediary Liability",
    9:  "Content Access",
    10: "Non-tariff Trade Policy",
    11: "Standards and Conformance",
    12: "E-commerce",
}

# Maximum number of split reference URL columns in the RDTII template export
_MAX_REF_COLS = 5

# ─── RDTII template column structure (matches the official Excel template) ────
# 13 columns total: 7 data + 5 ref URL slots + 1 note
RDTII_TEMPLATE_COLUMNS = [
    "Pillar_ID",
    "Indicator_ID",
    "Raw Score",
    "Act and/or practice",
    "Coverage",
    "Impact or comments on Acts or practices",
    "Timeframe",
    "References",
    "References_2",
    "References_3",
    "References_4",
    "References_5",
    "Note",
]

# ─── 9-column RDTII legacy export (§2.1) ─────────────────────────────────────
RDTII_COLUMNS = [
    "pillar_id",
    "indicator_id",
    "raw_score",
    "act_and_practice",
    "coverage",
    "impact_comments",
    "timeframe",
    "references",
    "note",
]

RDTII_LABELS = {
    "pillar_id": "Pillar_ID",
    "indicator_id": "Indicator_ID",
    "raw_score": "Raw Score",
    "act_and_practice": "Act and/or practice",
    "coverage": "Coverage",
    "impact_comments": "Impact or comments on Acts or practices",
    "timeframe": "Timeframe",
    "references": "References",
    "note": "Note",
}

# ─── Hackathon submission spec columns (§17) ─────────────────────────────────
SUBMISSION_COLUMNS = [
    "economy",
    "law_name",
    "law_number_ref",
    "last_amended",
    "indicator_id",
    "article_section",
    "discovery_tag",
    "location_ref",
    "verbatim_snippet",
    "mapping_rationale",
    "source_url",
    "confidence",
    "notes",
]

SUBMISSION_LABELS = {
    "economy": "Economy",
    "law_name": "Law Name",
    "law_number_ref": "Law Number / Ref",
    "last_amended": "Last Amended",
    "indicator_id": "Indicator ID",
    "article_section": "Article / Section",
    "discovery_tag": "Discovery Tag",
    "location_ref": "Location Reference",
    "verbatim_snippet": "Verbatim Snippet",
    "mapping_rationale": "Mapping Rationale",
    "source_url": "Source URL",
    "confidence": "Confidence",
    "notes": "Notes",
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _split_references(references: str | None) -> list[str]:
    """
    Split a newline- or semicolon-separated reference string into individual URLs.
    Returns a list of up to _MAX_REF_COLS non-empty strings.
    """
    if not references:
        return []
    # Support both newline- and semicolon-delimited references
    parts = []
    for line in references.splitlines():
        for part in line.split(";"):
            part = part.strip()
            if part:
                parts.append(part)
    return parts[:_MAX_REF_COLS]


def _pad_refs(refs: list[str]) -> tuple:
    """Pad a list of reference strings to exactly _MAX_REF_COLS elements."""
    padded = refs + [""] * (_MAX_REF_COLS - len(refs))
    return tuple(padded[:_MAX_REF_COLS])


def _results_to_records(indicator_results: list) -> list[dict[str, Any]]:
    """Convert ORM IndicatorResult objects to plain dicts for RDTII export."""
    records = []
    for r in sorted(indicator_results, key=lambda x: (x.pillar_id, x.indicator_id)):
        records.append({
            "pillar_id": r.pillar_id,
            "indicator_id": r.indicator_id,
            "raw_score": r.raw_score,
            "act_and_practice": r.act_and_practice or "—",
            "coverage": r.coverage or "N/A",
            "impact_comments": r.impact_comments or "—",
            "timeframe": r.timeframe or "—",
            "references": r.references or "—",
            "note": r.note or "—",
            # Audit fields (included in JSON, excluded from flat CSV)
            "confidence": r.confidence,
            "verbatim_quote": r.verbatim_quote,
            "article_citation": r.article_citation,
            "not_found": r.not_found,
            "prosecution_score": r.prosecution_score,
            "defense_score": r.defense_score,
            "arbiter_score": r.arbiter_score,
            "discovery_tag": r.discovery_tag or "NEW",
            "source_pdf_path": r.source_pdf_path,
            "location_ref": r.location_ref,
            "processing_time": r.processing_time,
            "mapping_rationale": r.mapping_rationale,
        })
    return records


def _results_to_rdtii_template_rows(indicator_results: list) -> list[dict[str, Any]]:
    """
    Convert ORM IndicatorResult objects to rows matching the official RDTII
    Excel template format:
      - Pillar section-header rows inserted at each pillar boundary
      - References split across up to 5 separate columns
      - Correct official column header names

    Returns a list of row dicts. Header rows have is_header=True.
    """
    sorted_results = sorted(indicator_results, key=lambda x: (x.pillar_id, x.indicator_id))
    rows = []
    current_pillar = None

    for r in sorted_results:
        pillar_id = r.pillar_id

        # Insert pillar section header when pillar changes
        if pillar_id != current_pillar:
            current_pillar = pillar_id
            pillar_name = PILLAR_NAMES.get(pillar_id, f"Pillar {pillar_id}")
            header_row = {
                "Pillar_ID": pillar_id,
                "Indicator_ID": pillar_name,
                "Raw Score": "",
                "Act and/or practice": "",
                "Coverage": "",
                "Impact or comments on Acts or practices": "",
                "Timeframe": "",
                "References": "",
                "References_2": "",
                "References_3": "",
                "References_4": "",
                "References_5": "",
                "Note": "",
                "_is_pillar_header": True,
            }
            rows.append(header_row)

        # Split references into up to 5 columns
        ref_parts = _split_references(r.references)
        ref1, ref2, ref3, ref4, ref5 = _pad_refs(ref_parts)

        data_row = {
            "Pillar_ID": pillar_id,
            "Indicator_ID": r.indicator_id,
            "Raw Score": r.raw_score if r.raw_score is not None else "",
            "Act and/or practice": r.act_and_practice or "",
            "Coverage": r.coverage or "",
            "Impact or comments on Acts or practices": r.impact_comments or "",
            "Timeframe": r.timeframe or "",
            "References": ref1,
            "References_2": ref2,
            "References_3": ref3,
            "References_4": ref4,
            "References_5": ref5,
            "Note": r.note or "",
            "_is_pillar_header": False,
        }
        rows.append(data_row)

    return rows


def _indicator_id_to_rdtii(pillar_id: int, indicator_id: str) -> str:
    """Convert '6.1' -> 'P6-I1', '12.4.1' -> 'P12-I4.1' using the RDTII format."""
    parts = indicator_id.split(".")
    suffix = ".".join(parts[1:]) if len(parts) > 1 else parts[0]
    return f"P{pillar_id}-I{suffix}"


def _results_to_submission_records(indicator_results: list, country: str) -> list[dict[str, Any]]:
    """Convert ORM IndicatorResult objects to submission-spec records (§17)."""
    records = []
    for r in sorted(indicator_results, key=lambda x: (x.pillar_id, x.indicator_id)):
        citation = r.article_citation or ""
        records.append({
            "economy": country,
            "law_name": r.act_and_practice or "—",
            "law_number_ref": r.law_number_ref or "—",
            "last_amended": r.timeframe or "—",
            "indicator_id": _indicator_id_to_rdtii(r.pillar_id, r.indicator_id),
            "article_section": citation or "—",
            "discovery_tag": r.discovery_tag or "NEW",
            "location_ref": r.location_ref or "",
            "verbatim_snippet": r.verbatim_quote or "—",
            "mapping_rationale": (r.mapping_rationale or r.impact_comments or "")[:300],
            "source_url": r.references or "—",
            "confidence": round(r.confidence, 2) if r.confidence is not None else "",
            "notes": r.note or "—",
        })
    return records


# ─── Public Export Functions ──────────────────────────────────────────────────

def export_json(indicator_results: list) -> str:
    """
    Export indicator results as a formatted JSON string.
    Includes all audit fields per submission spec §17.
    """
    records = _results_to_records(indicator_results)
    return json.dumps(records, indent=2, default=str)


def export_csv(indicator_results: list, format: str = "rdtii", country: str = "") -> io.BytesIO:
    """
    Export indicator results as CSV.
    Three format options:
      - "rdtii"      : Official RDTII template format — pillar header rows +
                       split reference columns (matches the official Excel template)
      - "rdtii_flat" : Legacy 9-column flat RDTII schema (§2.1)
      - "submission" : Hackathon submission spec columns (§17)

    Returns:
        BytesIO buffer containing UTF-8 CSV content.
    """
    if format == "rdtii":
        rows = _results_to_rdtii_template_rows(indicator_results)
        # Strip internal marker column before writing
        output_cols = RDTII_TEMPLATE_COLUMNS
        clean_rows = [{k: v for k, v in row.items() if k != "_is_pillar_header"} for row in rows]
        df = pd.DataFrame(clean_rows, columns=output_cols)
    elif format == "rdtii_flat":
        records = _results_to_records(indicator_results)
        df = pd.DataFrame(records, columns=RDTII_COLUMNS)
        df.rename(columns=RDTII_LABELS, inplace=True)
    else:
        records = _results_to_submission_records(indicator_results, country or "Unknown")
        df = pd.DataFrame(records, columns=SUBMISSION_COLUMNS)
        df.rename(columns=SUBMISSION_LABELS, inplace=True)

    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)
    return buf


def export_excel(indicator_results: list, country: str) -> io.BytesIO:
    """
    Export indicator results as an Excel (.xlsx) file.
    Produces three sheets:
      1. "RDTII_Template" — official RDTII template format with pillar header rows,
         split reference URL columns, and colour-coded scores.
      2. "RDTII_9col"    — legacy 9-column RDTII schema (§2.1)
      3. "Submission"    — Hackathon submission spec columns (§17)
    Includes formatting: frozen header row, auto column widths, colour-coded scores,
    styled pillar section headers.

    Args:
        indicator_results: List of IndicatorResult ORM objects.
        country: Country name — used in the submission sheet.

    Returns:
        BytesIO buffer containing XLSX content.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: RDTII Template format ──────────────────────────────────
        template_rows = _results_to_rdtii_template_rows(indicator_results)
        is_header_flags = [row.pop("_is_pillar_header", False) for row in template_rows]
        tmpl_df = pd.DataFrame(template_rows, columns=RDTII_TEMPLATE_COLUMNS)
        tmpl_df.to_excel(writer, sheet_name="RDTII_Template", index=False)

        # ── Sheet 2: Legacy 9-column RDTII ──────────────────────────────────
        rdtii_records = _results_to_records(indicator_results)
        rdtii_df = pd.DataFrame(rdtii_records, columns=RDTII_COLUMNS)
        rdtii_df.rename(columns=RDTII_LABELS, inplace=True)
        rdtii_df.to_excel(writer, sheet_name="RDTII_9col", index=False)

        # ── Sheet 3: Submission spec ─────────────────────────────────────────
        sub_records = _results_to_submission_records(indicator_results, country)
        sub_df = pd.DataFrame(sub_records, columns=SUBMISSION_COLUMNS)
        sub_df.rename(columns=SUBMISSION_LABELS, inplace=True)
        sub_df.to_excel(writer, sheet_name="Submission", index=False)

        # ── Apply formatting ─────────────────────────────────────────────────
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            # ── Format Sheet 1: RDTII_Template ──────────────────────────────
            ws_tmpl = writer.sheets["RDTII_Template"]

            # Header row styling (row 1)
            header_fill = PatternFill("solid", fgColor="1F3864")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws_tmpl[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws_tmpl.freeze_panes = "A2"

            # Style each data row
            pillar_header_fill = PatternFill("solid", fgColor="2E74B5")  # UN blue
            pillar_header_font = Font(bold=True, color="FFFFFF", size=11)

            score_col_idx = RDTII_TEMPLATE_COLUMNS.index("Raw Score") + 1
            note_col_idx = RDTII_TEMPLATE_COLUMNS.index("Note") + 1
            total_cols = len(RDTII_TEMPLATE_COLUMNS)

            for row_idx, is_pillar_hdr in enumerate(is_header_flags, start=2):
                if is_pillar_hdr:
                    # Style as a pillar section header spanning all columns
                    for col_idx in range(1, total_cols + 1):
                        cell = ws_tmpl.cell(row=row_idx, column=col_idx)
                        cell.fill = pillar_header_fill
                        cell.font = pillar_header_font
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    # Bold pillar name text in column B
                    ws_tmpl.cell(row=row_idx, column=2).font = Font(
                        bold=True, color="FFFFFF", size=12
                    )
                else:
                    # Colour-code Raw Score
                    score_cell = ws_tmpl.cell(row=row_idx, column=score_col_idx)
                    score_val = score_cell.value
                    if isinstance(score_val, (int, float)):
                        if score_val == 1.0:
                            score_cell.fill = PatternFill("solid", fgColor="FF4444")
                            score_cell.font = Font(bold=True, color="FFFFFF")
                        elif score_val == 0.5:
                            score_cell.fill = PatternFill("solid", fgColor="FFA500")
                            score_cell.font = Font(bold=True, color="FFFFFF")
                        elif score_val == 0.0:
                            score_cell.fill = PatternFill("solid", fgColor="44BB44")
                            score_cell.font = Font(bold=True, color="FFFFFF")

                    # Wrap text for long-text columns
                    for col_idx in range(1, total_cols + 1):
                        cell = ws_tmpl.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Auto-fit column widths (sample first 40 data rows)
            col_max_widths = {
                "Pillar_ID": 9,
                "Indicator_ID": 12,
                "Raw Score": 10,
                "Act and/or practice": 35,
                "Coverage": 16,
                "Impact or comments on Acts or practices": 55,
                "Timeframe": 30,
                "References": 50,
                "References_2": 50,
                "References_3": 50,
                "References_4": 50,
                "References_5": 50,
                "Note": 50,
            }
            for col_idx, col_name in enumerate(RDTII_TEMPLATE_COLUMNS, 1):
                ws_tmpl.column_dimensions[get_column_letter(col_idx)].width = col_max_widths.get(
                    col_name, 30
                )

            # Set row height for pillar header rows
            for row_idx, is_pillar_hdr in enumerate(is_header_flags, start=2):
                if is_pillar_hdr:
                    ws_tmpl.row_dimensions[row_idx].height = 22

            # ── Format Sheets 2 & 3 (legacy / submission) ───────────────────
            for sheet_name, df in [("RDTII_9col", rdtii_df), ("Submission", sub_df)]:
                ws = writer.sheets[sheet_name]

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.freeze_panes = "A2"

                if "Raw Score" in df.columns:
                    sc_idx = list(df.columns).index("Raw Score") + 1
                    for row_i in range(2, ws.max_row + 1):
                        sc_cell = ws.cell(row=row_i, column=sc_idx)
                        sc_val = sc_cell.value
                        if isinstance(sc_val, (int, float)):
                            if sc_val == 1.0:
                                sc_cell.fill = PatternFill("solid", fgColor="FF4444")
                                sc_cell.font = Font(bold=True, color="FFFFFF")
                            elif sc_val == 0.5:
                                sc_cell.fill = PatternFill("solid", fgColor="FFA500")
                                sc_cell.font = Font(bold=True, color="FFFFFF")
                            elif sc_val == 0.0:
                                sc_cell.fill = PatternFill("solid", fgColor="44BB44")
                                sc_cell.font = Font(bold=True, color="FFFFFF")

                for col_idx, col in enumerate(df.columns, 1):
                    max_len = max(
                        len(str(col)),
                        *[len(str(ws.cell(row=r, column=col_idx).value or ""))
                          for r in range(2, min(ws.max_row + 1, 20))]
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        except Exception:
            pass  # Formatting is optional — don't fail the export

    buf.seek(0)
    return buf
